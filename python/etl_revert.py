import json
import sqlite3
import csv
from datetime import date, datetime, timedelta, timezone
from io import StringIO
from pathlib import Path
from typing import Any

import requests


def excel_col_name(idx: int) -> str:
    n = idx + 1
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def load_config() -> dict[str, Any]:
    config_path = Path("python/config.json")
    if not config_path.exists():
        raise FileNotFoundError("Create python/config.json from python/config.example.json")
    return json.loads(config_path.read_text(encoding="utf-8"))


def normalize_symbol(symbol: str) -> str:
    s = symbol.strip().upper()
    if "BTC" in s:
        return "BTC"
    if "ETH" in s:
        return "ETH"
    return s


def parse_float(value: Any) -> float:
    s = str(value or "").strip().replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def fetch_prices() -> dict[str, float]:
    url = (
        "https://api.coingecko.com/api/v3/simple/price"
        "?ids=bitcoin,ethereum,tether,usd-coin&vs_currencies=usd"
    )
    res = requests.get(url, timeout=30)
    res.raise_for_status()
    data = res.json()
    return {
        "BTC": float(data["bitcoin"]["usd"]),
        "ETH": float(data["ethereum"]["usd"]),
        "USDT": float(data["tether"]["usd"]),
        "USDC": float(data["usd-coin"]["usd"]),
    }


def load_google_sheet_rows(sheet_id: str, gid: str) -> list[dict[str, Any]]:
    if not gid:
        return []
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    res = requests.get(url, timeout=30)
    res.raise_for_status()
    reader = csv.reader(StringIO(res.text))
    raw_rows = list(reader)
    return _sheet_csv_to_row_dicts(raw_rows)


def load_google_sheet_rows_by_name(sheet_id: str, sheet_name: str) -> list[dict[str, Any]]:
    if not sheet_id or not sheet_name:
        return []
    for name in (sheet_name, "Public portfolio", "Public Portfolio"):
        url = (
            f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq"
            f"?tqx=out:csv&sheet={requests.utils.quote(name)}"
        )
        res = requests.get(url, timeout=30)
        if not res.ok:
            continue
        raw_rows = list(csv.reader(StringIO(res.text)))
        rows = _sheet_csv_to_row_dicts(raw_rows)
        if rows:
            return rows
    return []


def _sheet_csv_to_row_dicts(raw_rows: list[list[str]]) -> list[dict[str, Any]]:
    if not raw_rows:
        return []
    headers = [str(x).strip() for x in raw_rows[0]]
    rows: list[dict[str, Any]] = []
    for row in raw_rows[1:]:
        cleaned: dict[str, Any] = {}
        for i, value in enumerate(row):
            col_letter = excel_col_name(i)
            cleaned[col_letter] = value.strip() if isinstance(value, str) else value
            if i < len(headers) and headers[i]:
                cleaned[headers[i]] = value.strip() if isinstance(value, str) else value
        rows.append(cleaned)
    return rows


def row_get(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if not key:
            continue
        if key in row and row[key] not in (None, ""):
            return row[key]
    return ""


def parse_fee_apy_pct(raw: Any) -> float:
    s = str(raw or "").strip()
    if not s or s in ("-", "—", "0"):
        return 0.0
    p = parse_float(s.replace("%", ""))
    if p <= 0:
        return 0.0
    if "%" in s or p >= 3:
        return min(p, 500.0)
    return min(p * 100.0, 500.0)


def liquidity_open_from_row(row: dict[str, Any]) -> bool:
    oc = str(row_get(row, "Open/Closed", "N")).strip().upper()
    if oc == "OPEN":
        return True
    if oc == "CLOSED":
        return False
    return not str(row_get(row, "Дата закрытия", "H", "X")).strip()


def upsert_market_prices(conn: sqlite3.Connection, as_of_date: str, prices: dict[str, float]) -> None:
    cur = conn.cursor()
    for symbol, px in prices.items():
        cur.execute(
            """
            INSERT INTO market_prices(as_of_date, symbol, price_usd, source)
            VALUES (?, ?, ?, 'coingecko')
            ON CONFLICT(as_of_date, symbol, source)
            DO UPDATE SET price_usd=excluded.price_usd, created_at=datetime('now')
            """,
            (as_of_date, symbol, px),
        )


def ensure_revert_history_table(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS revert_metrics_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_ts TEXT NOT NULL,
            as_of_date TEXT NOT NULL,
            liquidity_usd REAL NOT NULL DEFAULT 0,
            total_fee_usd REAL NOT NULL DEFAULT 0,
            earned_since_prev_usd REAL NOT NULL DEFAULT 0,
            elapsed_hours REAL NOT NULL DEFAULT 0,
            apr_annualized REAL NOT NULL DEFAULT 0
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS revert_daily_rollup (
            as_of_date TEXT PRIMARY KEY,
            run_ts TEXT NOT NULL,
            liquidity_usd REAL NOT NULL DEFAULT 0,
            total_fee_usd REAL NOT NULL DEFAULT 0,
            earned_vs_prev_day_usd REAL NOT NULL DEFAULT 0,
            elapsed_hours REAL NOT NULL DEFAULT 24,
            apr_annualized REAL NOT NULL DEFAULT 0
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS revert_positions_daily (
            as_of_date TEXT NOT NULL,
            position_id TEXT NOT NULL,
            pair TEXT NOT NULL DEFAULT '',
            chain TEXT NOT NULL DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 1,
            opened_at TEXT NOT NULL DEFAULT '',
            closed_at TEXT NOT NULL DEFAULT '',
            invested_usd REAL NOT NULL DEFAULT 0,
            fee_tier_pct REAL NOT NULL DEFAULT 0,
            total_fee_usd REAL NOT NULL DEFAULT 0,
            apr_pct REAL NOT NULL DEFAULT 0,
            link TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT (datetime('now')),
            PRIMARY KEY (as_of_date, position_id)
        )
        """
    )


def latest_history_row(conn: sqlite3.Connection) -> sqlite3.Row | None:
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    return cur.execute(
        """
        SELECT run_ts, total_fee_usd
        FROM revert_metrics_history
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()


def previous_day_rollup_row(conn: sqlite3.Connection, as_of_date: str) -> sqlite3.Row | None:
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        prev_day = (date.fromisoformat(as_of_date) - timedelta(days=1)).isoformat()
    except ValueError:
        prev_day = ""
    if not prev_day:
        return None
    row = cur.execute(
        """
        SELECT as_of_date, run_ts, total_fee_usd
        FROM revert_daily_rollup
        WHERE as_of_date = ?
        ORDER BY run_ts DESC
        LIMIT 1
        """,
        (prev_day,),
    ).fetchone()
    if row:
        return row
    return cur.execute(
        """
        SELECT as_of_date, run_ts, total_fee_usd
        FROM revert_metrics_history
        WHERE as_of_date = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (prev_day,),
    ).fetchone()


def ingest_revert(config: dict[str, Any]) -> None:
    db_path = config["db_path"]
    revert_source = config.get("revert_source", "excel")
    pools_cols = config["revert_pools_columns"]
    fees_cols = config["revert_fees_columns"]

    if not Path(db_path).exists():
        raise FileNotFoundError(f"DB not found: {db_path}. Run python python/init_db.py first.")
    excel_path = config.get("revert_excel_path", "")
    pools_sheet = config.get("revert_pools_sheet", "pools")
    fees_sheet = config.get("revert_fees_sheet", "fees")
    sheet_id = config.get("google_sheet_id", "")
    pools_gid = str(config.get("revert_pools_gid", "")).strip()
    fees_gid = str(config.get("revert_fees_gid", "")).strip()
    pools_sheet_name = str(config.get("revert_pools_sheet_name", "") or "").strip()
    fee_total_col = str(
        config.get("revert_fee_total_column", "")
        or config.get("revert_pools_columns", {}).get("fee_total_usd", "")
        or "Заработано комиссий всего, USD"
    ).strip()

    if revert_source == "excel" and (not excel_path or not Path(excel_path).exists()):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if revert_source == "google_sheets" and not sheet_id:
        raise ValueError("Set google_sheet_id in python/config.json")

    as_of_date = date.today().isoformat()
    prices = fetch_prices()

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        ensure_revert_history_table(conn)
        upsert_market_prices(conn, as_of_date, prices)

        pools_rows: list[dict[str, Any]]
        fees_rows_data: list[dict[str, Any]]
        if revert_source == "google_sheets":
            if pools_sheet_name:
                pools_rows = load_google_sheet_rows_by_name(sheet_id, pools_sheet_name)
            else:
                pools_rows = load_google_sheet_rows(sheet_id, pools_gid)
            fees_rows_data = load_google_sheet_rows(sheet_id, fees_gid) if fees_gid else []
            if not pools_rows:
                raise ValueError("Pools Google Sheet is empty or not accessible")
        else:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            if pools_sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{pools_sheet}' not found")
            ws_pools = wb[pools_sheet]
            rows = list(ws_pools.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Pools sheet is empty")
            headers = [str(x).strip() if x is not None else "" for x in list(rows[0])]
            data_rows = rows[1:]
            pools_rows = []
            for row in data_rows:
                pools_rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]})

            fees_rows_data = []
            if fees_sheet in wb.sheetnames:
                ws_fees = wb[fees_sheet]
                fee_rows = list(ws_fees.iter_rows(values_only=True))
                if fee_rows:
                    fee_headers = [str(x).strip() if x is not None else "" for x in list(fee_rows[0])]
                    for row in fee_rows[1:]:
                        fees_rows_data.append(
                            {fee_headers[i]: row[i] for i in range(min(len(fee_headers), len(row))) if fee_headers[i]}
                        )

        pair_mode = all(
            key in pools_cols
            for key in ["token0_symbol", "token1_symbol", "token0_amount", "token1_amount"]
        )

        required_pool_cols = [pools_cols["position_id"]]
        if pair_mode:
            required_pool_cols.extend(
                [pools_cols["token0_symbol"], pools_cols["token1_symbol"], pools_cols["token0_amount"], pools_cols["token1_amount"]]
            )
        else:
            required_pool_cols.extend([pools_cols["symbol"], pools_cols["amount"]])
        available_pool_headers = list(pools_rows[0].keys())
        for c in required_pool_cols:
            if c not in pools_rows[0]:
                raise ValueError(f"Pools column not found: {c}. Available: {available_pool_headers}")

        cur.execute("DELETE FROM revert_positions WHERE as_of_date = ?", (as_of_date,))
        cur.execute("DELETE FROM revert_positions_daily WHERE as_of_date = ?", (as_of_date,))
        liquidity_usd = 0.0

        for row in pools_rows:
            position_id = row.get(pools_cols["position_id"])
            if position_id is None or str(position_id).strip() == "":
                continue

            token_items: list[tuple[str, Any]] = []
            if pair_mode:
                token_items = [
                    (str(row.get(pools_cols["token0_symbol"], "")), row.get(pools_cols["token0_amount"])),
                    (str(row.get(pools_cols["token1_symbol"], "")), row.get(pools_cols["token1_amount"])),
                ]
            else:
                token_items = [(str(row.get(pools_cols["symbol"], "")), row.get(pools_cols["amount"]))]

            for symbol_raw, amount_raw in token_items:
                if symbol_raw is None or amount_raw is None or str(symbol_raw).strip() == "":
                    continue
                try:
                    amount = float(amount_raw)
                except (TypeError, ValueError):
                    continue
                symbol = normalize_symbol(str(symbol_raw))
                px = prices.get(symbol, 0.0)
                usd_value = amount * px
                liquidity_usd += usd_value

                cur.execute(
                    """
                    INSERT INTO revert_positions(as_of_date, position_id, symbol, amount, usd_value)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (as_of_date, str(position_id), symbol, amount, usd_value),
                )

        if liquidity_usd <= 0:
            for row in pools_rows:
                if not liquidity_open_from_row(row):
                    continue
                pos_val = parse_float(
                    row_get(
                        row,
                        pools_cols.get("position_value_usd", ""),
                        "Стоимость позиции, USD",
                        "Внесено, USD",
                    )
                )
                if pos_val > 0:
                    liquidity_usd += pos_val

        latest_fee_total_usd = 0.0
        # Preferred source: AG in pools sheet is cumulative total fee in USD by position.
        fee_total_from_ag = 0.0
        fee_ag_count = 0
        for row in pools_rows:
            v = parse_float(row_get(row, fee_total_col, "AG", "Заработано комиссий всего, USD"))
            if v > 0:
                fee_total_from_ag += v
                fee_ag_count += 1

        fee_source_rows = fees_rows_data if fees_rows_data else pools_rows
        if fee_source_rows:
            if fee_ag_count > 0:
                latest_fee_total_usd = fee_total_from_ag
            elif fees_cols.get("token0_fee") and fees_cols.get("token1_fee"):
                token0_fee_col = fees_cols["token0_fee"]
                token1_fee_col = fees_cols["token1_fee"]
                token0_symbol_col = fees_cols.get("token0_symbol", pools_cols.get("token0_symbol"))
                token1_symbol_col = fees_cols.get("token1_symbol", pools_cols.get("token1_symbol"))
                for row in fee_source_rows:
                    fee0 = row.get(token0_fee_col)
                    fee1 = row.get(token1_fee_col)
                    sym0 = str(row.get(token0_symbol_col, "")) if token0_symbol_col else ""
                    sym1 = str(row.get(token1_symbol_col, "")) if token1_symbol_col else ""
                    fee0_num = parse_float(fee0)
                    fee1_num = parse_float(fee1)
                    latest_fee_total_usd += (
                        fee0_num * prices.get(normalize_symbol(sym0), 0.0)
                        + fee1_num * prices.get(normalize_symbol(sym1), 0.0)
                    )
            elif fees_cols.get("daily_fee_income_usd") and fees_cols.get("date"):
                if fees_cols["date"] not in fees_rows_data[0] or fees_cols["daily_fee_income_usd"] not in fees_rows_data[0]:
                    available_fee_headers = list(fees_rows_data[0].keys())
                    raise ValueError(
                        f"Fees columns not found: {fees_cols}. Available: {available_fee_headers}"
                    )
                latest_row_date = None
                for row in fees_rows_data:
                    row_date = row.get(fees_cols["date"])
                    fee_raw = row.get(fees_cols["daily_fee_income_usd"])
                    if row_date is None or fee_raw is None:
                        continue
                    latest_row_date = str(row_date)
                    latest_fee_total_usd = parse_float(fee_raw)
                if latest_row_date is None:
                    latest_fee_total_usd = 0.0

        # Annualized APR based on fee growth since previous day checkpoint.
        run_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        prev = previous_day_rollup_row(conn, as_of_date)
        earned_since_prev = 0.0
        elapsed_hours = 24.0
        apr_annualized = 0.0
        if prev:
            try:
                prev_ts_raw = str(prev["run_ts"])
                try:
                    prev_ts = datetime.strptime(prev_ts_raw, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                except ValueError:
                    prev_ts = datetime.strptime(prev_ts_raw[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                now_ts = datetime.strptime(run_ts, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                elapsed_hours = max((now_ts - prev_ts).total_seconds() / 3600.0, 0.0)
            except ValueError:
                elapsed_hours = 24.0
            if elapsed_hours <= 0:
                elapsed_hours = 24.0
            prev_total = float(prev["total_fee_usd"] or 0.0)
            earned_since_prev = max(0.0, latest_fee_total_usd - prev_total)
            if elapsed_hours > 0 and liquidity_usd > 0 and earned_since_prev > 0:
                # Минимум ~20ч между прогонами: иначе при двух запусках в день APR раздувается до 100%+.
                eff_hours = max(elapsed_hours, 20.0)
                apr_annualized = (earned_since_prev / liquidity_usd) * (24.0 * 365.0 / eff_hours) * 100.0
                apr_annualized = min(apr_annualized, 500.0)

        cur.execute(
            """
            INSERT INTO daily_metrics(
                as_of_date, liquidity_usd, daily_fee_income_usd, updated_at, source_revert
            ) VALUES (?, ?, ?, datetime('now'), 'excel')
            ON CONFLICT(as_of_date) DO UPDATE SET
                liquidity_usd=excluded.liquidity_usd,
                daily_fee_income_usd=excluded.daily_fee_income_usd,
                updated_at=datetime('now'),
                source_revert='excel'
            """,
            (as_of_date, liquidity_usd, earned_since_prev),
        )
        cur.execute(
            """
            INSERT INTO revert_metrics_history(
                run_ts, as_of_date, liquidity_usd, total_fee_usd, earned_since_prev_usd, elapsed_hours, apr_annualized
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (run_ts, as_of_date, liquidity_usd, latest_fee_total_usd, earned_since_prev, elapsed_hours, apr_annualized),
        )
        cur.execute(
            """
            INSERT INTO revert_daily_rollup(
                as_of_date, run_ts, liquidity_usd, total_fee_usd, earned_vs_prev_day_usd, elapsed_hours, apr_annualized
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(as_of_date) DO UPDATE SET
                run_ts=excluded.run_ts,
                liquidity_usd=excluded.liquidity_usd,
                total_fee_usd=excluded.total_fee_usd,
                earned_vs_prev_day_usd=excluded.earned_vs_prev_day_usd,
                elapsed_hours=excluded.elapsed_hours,
                apr_annualized=excluded.apr_annualized
            """,
            (as_of_date, run_ts, liquidity_usd, latest_fee_total_usd, earned_since_prev, elapsed_hours, apr_annualized),
        )

        cur.execute(
            """
            INSERT INTO ingestion_runs(as_of_date, pipeline, status, message)
            VALUES (?, 'revert_excel', 'success', ?)
            """,
            (
                as_of_date,
                (
                    f"liquidity_usd={liquidity_usd:.2f}; total_fee={latest_fee_total_usd:.2f}; "
                    f"earned_since_prev={earned_since_prev:.2f}; elapsed_h={elapsed_hours:.2f}; apr={apr_annualized:.2f}"
                ),
            ),
        )

        for r in pools_rows:
            position_id = str(row_get(r, pools_cols["position_id"], "NFT tokenId", "BT")).strip()
            if not position_id:
                continue
            token0 = str(row_get(r, pools_cols.get("token0_symbol", ""), "Токен 0", "AK") or "").strip()
            token1 = str(row_get(r, pools_cols.get("token1_symbol", ""), "Токен 1", "AL") or "").strip()
            pair = f"{token0} / {token1}".strip(" /")
            chain = str(row_get(r, "Сеть", "BH") or "").strip()
            opened_at = str(row_get(r, "Дата открытия", "W") or "").strip()
            closed_at = str(row_get(r, "Дата закрытия", "X") or "").strip()
            is_active = 1 if liquidity_open_from_row(r) else 0
            invested_usd = parse_float(
                row_get(r, pools_cols.get("invested_usd", ""), "Внесено, USD", "AA")
            )
            fee_tier_raw = row_get(r, "Fee tier (%)", "Fee tier", "BJ")
            fee_tier_pct = parse_float(str(fee_tier_raw).replace("%", ""))
            if "%" in str(fee_tier_raw) and fee_tier_pct >= 1000:
                fee_tier_pct /= 1000000
            elif "%" in str(fee_tier_raw) and fee_tier_pct >= 1:
                fee_tier_pct /= 100
            elif fee_tier_pct >= 50:
                fee_tier_pct /= 10000
            total_fee_usd = parse_float(row_get(r, fee_total_col, "AG", "Заработано комиссий всего, USD"))
            apr_pct = parse_fee_apy_pct(row_get(r, "Fee APY", "S")) if is_active else 0.0
            link = str(row_get(r, "Ссылка на позицию", "I") or "").strip()
            cur.execute(
                """
                INSERT INTO revert_positions_daily(
                    as_of_date, position_id, pair, chain, is_active, opened_at, closed_at,
                    invested_usd, fee_tier_pct, total_fee_usd, apr_pct, link, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(as_of_date, position_id) DO UPDATE SET
                    pair=excluded.pair,
                    chain=excluded.chain,
                    is_active=excluded.is_active,
                    opened_at=excluded.opened_at,
                    closed_at=excluded.closed_at,
                    invested_usd=excluded.invested_usd,
                    fee_tier_pct=excluded.fee_tier_pct,
                    total_fee_usd=excluded.total_fee_usd,
                    apr_pct=excluded.apr_pct,
                    link=excluded.link,
                    updated_at=datetime('now')
                """,
                (
                    as_of_date,
                    position_id,
                    pair,
                    chain,
                    is_active,
                    opened_at,
                    closed_at,
                    invested_usd,
                    fee_tier_pct,
                    total_fee_usd,
                    apr_pct,
                    link,
                ),
            )
        conn.commit()
        print(
            f"[OK] Revert ingested for {as_of_date}: liquidity_usd={liquidity_usd:.2f}, "
            f"total_fee={latest_fee_total_usd:.2f}, earned_since_prev={earned_since_prev:.2f}, "
            f"elapsed_h={elapsed_hours:.2f}, apr={apr_annualized:.2f}"
        )
    except Exception as exc:
        cur.execute(
            """
            INSERT INTO ingestion_runs(as_of_date, pipeline, status, message)
            VALUES (?, 'revert_excel', 'error', ?)
            """,
            (as_of_date, str(exc)),
        )
        conn.commit()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    cfg = load_config()
    ingest_revert(cfg)
